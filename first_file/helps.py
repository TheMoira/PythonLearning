from titles import *
from pathlib import Path
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def modify_excel_column_value(with_chart = 1):
    file0 = input("Enter the name of xlsx file: ")
    mod_file = f"{file0}v2.xlsx"
        #add exception od nonexisting file
    sheet_choice = input("Enter the sheet number: ")
    column_choice = int(input("Which column should be modified? "))
        #operation = input("Enter the operation, e.g. *3: ")
    number = int(input("Value should be multiplied by: "))

    wb = xl.load_workbook(f"{file0}.xlsx")
    sheet = wb[f"Arkusz{sheet_choice}"]

    for row in range(2, sheet.max_row + 1):
        current_cell_value = sheet.cell(row,column_choice).value
        new_cell_value = current_cell_value * number
        sheet.cell(row, column_choice).value = new_cell_value

    #wb.save(mod_file)

    if with_chart:
        chart_place = input("Where should the chart be? (Column symbol): ")
        values = Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=column_choice,
            max_col=column_choice
        )

        chart = BarChart()
        chart.add_data(values)
        sheet.add_chart(chart, chart_place.upper())
        wb.save(mod_file)

    print("The file has been modified.")
    print("The modified file can be found here: ")
    print(Path(mod_file))


